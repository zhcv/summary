import re
import openpyxl
from copy import copy

def cell_style(cell):
    alignment = copy(cell.alignment)    # 对齐样式
    border = copy(cell.border)          # 边框样式
    fill = copy(cell.fill)              # 填充样式
    font = copy(cell.font)              # 字体样式
    return alignment, border, fill, font

wb = openpyxl.load_workbook('工资条.xlsx') 
wb.copy_worksheet(wb['工资条'])
ws = wb.worksheets[-1]
ws.title = 'final_工资条'

# 获取每一列的值，拼接在一个列表中
cells_rows = [[cell for cell in row] for row in ws.rows]

# 获取标题
header = [cell.value for cell in cells_rows[0]]

# 获取标题行中，每个单元格中的各种样式
alignment, border, fill, font = cell_style(cell=cells_rows[0][0])

for i, _ in enumerate(cells_rows[1:]):
    if i > 0:
        index = i*3
        # 每读取一行，就在下方插入两行
        ws.insert_rows(idx=index, amount=2)
        # 写表头
        for j, v in enumerate(header):
            r, c = index+1, j+1
            cell = ws.cell(row=r, column=c)
            cell.value = v
            cell.alignment = alignment
            cell.font = font
            cell.border = border
            cell.fill = fill
            # 更新后面的公式
            if cell.coordinate[:1] in ('H', 'J'):
                cell = ws.cell(row=r+1, column=c)
                cell.value = re.sub('\d+', str(r+1), cell.value)
# 整个代码写完后，一定要记得保存               
wb.save('工资条.xlsx')    
